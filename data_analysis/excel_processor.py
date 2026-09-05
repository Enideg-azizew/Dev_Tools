#!/usr/bin/env python3
"""
Excel processing utilities.
Combines excel_.py and _from_excel.py.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Any, Optional

class ExcelProcessor:
    """Process Excel files with student grade data."""
    
    def __init__(self, file_path: Optional[str] = None):
        self.file_path = Path(file_path) if file_path else None
        self.workbook = None
    
    def load(self, file_path: str) -> None:
        """Load an Excel file."""
        self.file_path = Path(file_path)
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
    
    def read_sheet(self, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """Read a sheet as nested list."""
        if not self.workbook:
            raise ValueError("No workbook loaded")
        
        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        data = []
        
        for row in sheet.iter_rows(values_only=True):
            processed = []
            for cell in row:
                if cell is None:
                    processed.append("")
                elif isinstance(cell, (int, float)):
                    processed.append(cell)
                else:
                    processed.append(str(cell))
            data.append(processed)
        
        return data
    
    def save(self, data: List[List[Any]], filename: str = "output.xlsx") -> None:
        """Save data to Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_length + 2, 30)
        
        wb.save(filename)
        print(f"Saved to {filename}")

def main():
    import sys
    if len(sys.argv) < 2:
        print("Usage: python excel_processor.py <input_file> [output_file]")
        sys.exit(1)
    
    processor = ExcelProcessor()
    processor.load(sys.argv[1])
    data = processor.read_sheet()
    
    output = sys.argv[2] if len(sys.argv) > 2 else "processed.xlsx"
    processor.save(data, output)

if __name__ == "__main__":
    main()
